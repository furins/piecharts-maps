from openpyxl import load_workbook
import functools
import operator

import utility
from dataplot import plot_spicchi_shp, plot_spicchi_dxf, plot_diametri_shp, plot_diametri_dxf
import tkinter as tk


def carica_cpm(parent):
    def diametro(valore, classi_individuate, diam_min, inc_diam):
        for i, estremi in enumerate(classi_individuate):
            if estremi[0] < valore <= estremi[1]:
                return diam_min + i * inc_diam
        return diam_min

    utility.svuota(parent)
    utility.scrivi(parent, 'Rev. Software 2.0.8')
    utility.scrivi(parent, 'Elaborazione CPM iniziata... attendere')
    utility.avanzamento(parent, 1)

    diametro_minimo = float(parent.diametro_text.get('0.0', tk.END).strip())
    incremento_diametro = float(parent.step_cpm_text.get('0.0', tk.END).strip())

    # carica le classi
    dati = load_workbook(parent.file_text.get('0.0', tk.END).strip(), read_only=True, data_only=True)
    sheet_classi = dati['Classificazione']
    sheet_classi.calculate_dimension()
    colonna_limite_superiore = [x.value.strip() for x in sheet_classi[1] if x.value is not None].index(
        'Limite superiore') + 1
    limiti_classi = [0.0] + [float(x[colonna_limite_superiore].value)
                             for x in sheet_classi[2:sheet_classi.max_row]
                             if x[colonna_limite_superiore].value is not None]
    classi = tuple(zip(limiti_classi, limiti_classi[1:]))

    da_plottare = {}
    # importa le tavole con le posizioni, le labels, il dato da classificare, il diametro
    for nome_foglio_di_calcolo in dati.get_sheet_names():
        if not nome_foglio_di_calcolo.strip() == 'Classificazione':
            foglio = dati[nome_foglio_di_calcolo]
            foglio.calculate_dimension()

            colonne = [x.value.strip() for x in foglio[1] if x.value is not None]
            colonna_id = [n for n, r in enumerate(colonne) if r.strip().upper().startswith('ID')][0]
            colonna_x = [n for n, r in enumerate(colonne) if r.strip().upper().startswith('X')][0]
            colonna_y = [n for n, r in enumerate(colonne) if r.strip().upper().startswith('Y')][0]
            colonna_val = [n for n, r in enumerate(colonne) if r.strip().upper().startswith('SPESSORE SURNATANTE')][0]

            valori = []
            for row in foglio[2:foglio.max_row]:
                if row[0].value is not None:
                    valori += [{
                        'ID': row[colonna_id].value.strip(),
                        'X': float(row[colonna_x].value),
                        'Y': float(row[colonna_y].value),
                        'V': float(row[colonna_val].value),
                        'D': diametro(float(row[colonna_val].value), classi, diametro_minimo, incremento_diametro)
                    }]
                else:
                    # mi fermo alla prima riga vuota
                    break

            da_plottare[nome_foglio_di_calcolo] = valori

    # plotta shp
    # plotta dxf
    print(str(da_plottare))
    plot_diametri_shp(da_plottare, parent.esportazione_text.get('0.0', tk.END).strip())
    utility.avanzamento(parent, 50)
    plot_diametri_dxf(da_plottare, parent.esportazione_text.get('0.0', tk.END).strip(), diametro_minimo)
    utility.avanzamento(parent, 100)
    utility.scrivi(parent, 'Operazione completata!')
    pass


def carica_dati(parent):
    utility.svuota(parent)
    utility.scrivi(parent, 'Rev. Software 2.0.8')
    utility.scrivi(parent, 'Elaborazione PIE iniziata... attendere')
    utility.avanzamento(parent, 1)
    # prima acquisisco tutti i raggruppamenti
    inquinanti = load_workbook(parent.configurazione_text.get('0.0', tk.END).strip(), read_only=True)
    ws_inquinanti = inquinanti['classificazione']
    ws_inquinanti.calculate_dimension()

    inquinanti_noti = {x[0].value.strip(): (x[1].value.strip(), x[2].value)
                       for x in ws_inquinanti
                       if x[0].value is not None}

    wb = load_workbook(parent.file_text.get('0.0', tk.END).strip(), read_only=True)
    contatore = 0
    totale = len([x for x in wb.get_sheet_names() if x != 'RIEPILOGO'])
    for nome_foglio_di_calcolo in wb.get_sheet_names():
        if not nome_foglio_di_calcolo.strip() == 'RIEPILOGO':
            ws = wb[nome_foglio_di_calcolo]
            ws.calculate_dimension()
            famiglie = {x.value.strip(): inquinanti_noti[x.value.strip()][0]
                        for x in ws[1]
                        if x.value not in (None, 'ID', 'X', 'Y', 'PARAMETRO')}

            # le famiglie sono raggruppate sulla base del colore di sfondo della cella

            columns = [x.value for x in ws[1] if x.value is not None]

            raggruppamenti = {}
            for i, elemento in enumerate(columns[2:len(columns)]):
                elemento = elemento.strip()
                if elemento in famiglie.keys():
                    famiglia = str(famiglie[elemento])
                    try:
                        raggruppamenti[famiglia]['elementi'] += [i - 1]
                        raggruppamenti[famiglia]['soglie'] += [inquinanti_noti[elemento][1]]
                    except KeyError:
                        raggruppamenti[famiglia] = {'nome': inquinanti_noti[elemento][0]}
                        raggruppamenti[famiglia]['elementi'] = [i - 1]
                        raggruppamenti[famiglia]['soglie'] = [inquinanti_noti[elemento][1]]

            # -- fine  raggruppamenti per il foglio --

            # -- carico dati --
            dati = {}
            for row in ws[2:ws.max_row]:
                if row[0].value is not None:
                    dati[row[0].value] = dict(zip(columns[0:len(columns)], [x.value for x in row[0:len(columns)]]))
                else:
                    # mi fermo alla prima riga vuota
                    break
            # -- fine caricamento dati --

            dati_nuovi = {
                "colonne": columns[3:len(columns)],
                "raggruppamenti": raggruppamenti,
                "dati": dati
            }

            superamenti = prepara_superamenti(dati_nuovi, parent.inclusivo)

            for layout in superamenti.keys():
                # print('-' * 80)
                # print(nome_foglio_di_calcolo)
                # print('-' * 80)
                # pprint.pprint(dati_nuovi, depth=5, width=50)
                # pprint.pprint(superamenti[layout], depth=5, width=50)
                plot_spicchi_shp(nome_foglio_di_calcolo, parent.esportazione_text.get('0.0', tk.END).strip(), layout,
                                 superamenti[layout], dati_nuovi, int(parent.diametro_text.get('0.0', tk.END).strip()))
                plot_spicchi_dxf(nome_foglio_di_calcolo, parent.esportazione_text.get('0.0', tk.END).strip(), layout,
                                 superamenti[layout], dati_nuovi, int(parent.diametro_text.get('0.0', tk.END).strip()))
                # print('-'*80)
            contatore += 1
            utility.avanzamento(parent, 100 * contatore / totale)
    utility.scrivi(parent, 'Operazione completata!')


def prepara_superamenti(dati, inclusivo=1):
    # calcolo superamenti

    superamenti = {}

    for layout in dati["raggruppamenti"].keys():
        nome = str(layout) + '_' + dati["raggruppamenti"][layout]['nome']

        superamenti[nome] = {}
        for campione in dati["dati"].keys():
            superamenti[nome][campione] = {
                "X": dati["dati"][campione]["X"],
                "Y": dati["dati"][campione]["Y"],
                "SPICCHI": [False] * len(dati["raggruppamenti"][layout]["elementi"])
            }
            for i, elemento in enumerate(dati["raggruppamenti"][layout]["elementi"]):
                inquinante = dati["colonne"][elemento]
                try:
                    if (inclusivo == 1):
                        superamenti[nome][campione]["SPICCHI"][i] = float(dati["dati"][campione][inquinante]) >= dati["raggruppamenti"][layout]["soglie"][i]
                    else:
                        superamenti[nome][campione]["SPICCHI"][i] = float(dati["dati"][campione][inquinante]) > dati["raggruppamenti"][layout]["soglie"][i]
                except ValueError:
                    # il valore nella tabella non è un numero
                    superamenti[nome][campione]["SPICCHI"][i] = False
            superamenti[nome][campione]["HA_SUPERAMENTI"] = functools.reduce(operator.__or__,
                                                                             superamenti[nome][campione]["SPICCHI"])
    return superamenti
